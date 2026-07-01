"""导出工具测试 — CSV / Excel"""

import csv
import os
import tempfile

from openpyxl import load_workbook

from ui_pyside6.views.export_helper import export_to_csv, export_to_excel

HEADERS = ["物品名称", "数量", "价格", "备注"]
ROWS = [
    ["三钛合金", 10000, 5.5, "Jita买"],
    ["类银超金属", 5000, 9.2, "Jita买"],
    ["渡鸦级", 1, 52000000, "造船"],
]


class TestExportCsv:
    def test_export_csv_creates_file(self):
        """写入临时 CSV 文件，验证文件存在且非空"""
        with tempfile.NamedTemporaryFile(suffix=".csv", delete=False) as f:
            path = f.name
        try:
            export_to_csv(HEADERS, ROWS, path)
            assert os.path.exists(path)
            assert os.path.getsize(path) > 0
        finally:
            os.unlink(path)

    def test_export_csv_utf8_bom(self):
        """验证文件以 UTF-8 BOM (\\xef\\xbb\\xbf) 开头"""
        with tempfile.NamedTemporaryFile(suffix=".csv", delete=False) as f:
            path = f.name
        try:
            export_to_csv(HEADERS, ROWS, path)
            with open(path, "rb") as f:
                raw = f.read()
            assert raw[:3] == b"\xef\xbb\xbf", "缺少 UTF-8 BOM"
        finally:
            os.unlink(path)

    def test_export_csv_content(self):
        """验证内容和结构：首行为表头，后续为数据行"""
        with tempfile.NamedTemporaryFile(suffix=".csv", delete=False) as f:
            path = f.name
        try:
            export_to_csv(HEADERS, ROWS, path)
            with open(path, encoding="utf-8-sig") as f:
                reader = csv.reader(f)
                rows = list(reader)
            assert len(rows) == 4  # 1 header + 3 data
            assert rows[0] == HEADERS
            assert rows[1] == ["三钛合金", "10000", "5.5", "Jita买"]
            assert rows[2][0] == "类银超金属"
        finally:
            os.unlink(path)


class TestExportExcel:
    def test_export_excel_creates_file(self):
        """写入临时 .xlsx，验证文件存在且非空"""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = f.name
        try:
            export_to_excel(HEADERS, ROWS, path)
            assert os.path.exists(path)
            assert os.path.getsize(path) > 0
        finally:
            os.unlink(path)

    def test_export_excel_has_sheet_and_data(self):
        """验证 workbook 有默认 sheet 且包含数据"""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = f.name
        try:
            export_to_excel(HEADERS, ROWS, path)
            wb = load_workbook(path)
            ws = wb.active
            assert ws is not None
            assert ws.title == "Sheet1"
            # 表头
            assert [c.value for c in ws[1]] == HEADERS
            # 数据行
            assert ws.cell(2, 1).value == "三钛合金"
            assert ws.cell(3, 1).value == "类银超金属"
            assert ws.cell(4, 1).value == "渡鸦级"
            # 数值类型保留
            assert ws.cell(2, 2).value == 10000
        finally:
            os.unlink(path)

    def test_export_excel_empty_rows(self):
        """无数据时仅表头"""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = f.name
        try:
            export_to_excel(HEADERS, [], path)
            wb = load_workbook(path)
            ws = wb.active
            assert [c.value for c in ws[1]] == HEADERS
            assert ws.max_row == 1  # 只有表头
        finally:
            os.unlink(path)


class TestExportCsvEmpty:
    """CSV 空数据导出"""

    def test_export_csv_empty_rows(self):
        """无数据时仅表头"""
        import csv
        import os
        import tempfile

        from ui_pyside6.views.export_helper import export_to_csv

        headers = ["物品名称", "数量"]
        with tempfile.NamedTemporaryFile(suffix=".csv", delete=False) as f:
            path = f.name
        try:
            export_to_csv(headers, [], path)
            with open(path, encoding="utf-8-sig") as f:
                reader = csv.reader(f)
                rows = list(reader)
            assert len(rows) == 1  # 只有表头
            assert rows[0] == headers
        finally:
            os.unlink(path)

    def test_export_csv_none_values(self):
        """None 值应转为空字符串"""
        import csv
        import os
        import tempfile

        from ui_pyside6.views.export_helper import export_to_csv

        headers = ["A", "B"]
        rows = [[None, "hello"], ["world", None]]
        with tempfile.NamedTemporaryFile(suffix=".csv", delete=False) as f:
            path = f.name
        try:
            export_to_csv(headers, rows, path)
            with open(path, encoding="utf-8-sig") as f:
                reader = csv.reader(f)
                result = list(reader)
            assert len(result) == 3  # header + 2 data
            # None → empty string in CSV
            assert result[1][0] == ""  # None → ""
            assert result[1][1] == "hello"
            assert result[2][0] == "world"
            assert result[2][1] == ""  # None → ""
        finally:
            os.unlink(path)


class TestExportToNonexistentDir:
    """导出到不存在的目录应抛出 FileNotFoundError"""

    def test_csv_nonexistent_dir_raises(self):
        """export_to_csv 传入不存在的目录应抛 FileNotFoundError"""
        import os
        import tempfile

        from ui_pyside6.views.export_helper import export_to_csv

        headers = ["A"]
        rows = [["1"]]
        bad_path = os.path.join(tempfile.gettempdir(), "eve_test_nonexistent", "out.csv")
        assert not os.path.exists(os.path.dirname(bad_path))
        try:
            export_to_csv(headers, rows, bad_path)
            assert False, "应抛出 FileNotFoundError"
        except FileNotFoundError:
            pass

    def test_excel_nonexistent_dir_raises(self):
        """export_to_excel 传入不存在的目录应抛 FileNotFoundError"""
        import os
        import tempfile

        from ui_pyside6.views.export_helper import export_to_excel

        headers = ["A"]
        rows = [["1"]]
        bad_path = os.path.join(tempfile.gettempdir(), "eve_test_nonexistent", "out.xlsx")
        assert not os.path.exists(os.path.dirname(bad_path))
        try:
            export_to_excel(headers, rows, bad_path)
            assert False, "应抛出 FileNotFoundError"
        except FileNotFoundError:
            pass

    def test_csv_nonexistent_dir_deep_path(self):
        """深层不存在的目录也应抛出 FileNotFoundError"""
        import os
        import tempfile

        from ui_pyside6.views.export_helper import export_to_csv

        headers = ["A"]
        rows = [["1"]]
        bad_path = os.path.join(tempfile.gettempdir(), "eve_test_nonexistent", "sub", "nested", "out.csv")
        try:
            export_to_csv(headers, rows, bad_path)
            assert False, "应抛出 FileNotFoundError"
        except FileNotFoundError:
            pass
